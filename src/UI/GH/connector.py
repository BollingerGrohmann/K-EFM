import System
import Rhino
import Grasshopper
import Eto.Forms as forms
import Eto.Drawing as drawing
import scriptcontext as sc
import json
import urllib.parse  # Import to handle URL encoding
import openpyxl

from Grasshopper.Kernel.Data import GH_Path
from Grasshopper import DataTree
from System import Object

class Element:
    def __init__(self, name, geometries, thickness):
        self.name = name
        self.geometries = geometries
        self.thickness = thickness

class MyComponent:
    
    #slider_values = {}
    @staticmethod
    def RunScript(gh_doc, component, get_inputs, path):
        # Define a standalone function to recompute Grasshopper solution
        def schedule_recompute():
            # Get the current Grasshopper document
            gh_doc = Grasshopper.Instances.ActiveCanvas.Document
            if gh_doc is not None and component is not None:
                # Schedule a solution to recompute the Python component
                gh_doc.ScheduleSolution(1, lambda doc: component.ExpireSolution(True))

        # Define a function to select geometries from Rhino
        def get_surfaces():
            # Allow the user to select multiple surfaces
            go = Rhino.Input.Custom.GetObject()
            go.SetCommandPrompt("Select surfaces")
            go.GeometryFilter = Rhino.DocObjects.ObjectType.Surface
            go.EnablePreSelect(False, True)
            go.EnablePostSelect(True)
            go.GetMultiple(1, 0)  # Require at least one surface

            if go.CommandResult() != Rhino.Commands.Result.Success:
                return None

            # Collect selected surfaces
            surfaces = []
            for obj_ref in go.Objects():
                surface = obj_ref.Surface()
                if surface:
                    surfaces.append(surface)
            
            return surfaces
        
        
        def refresh_sender():
            for obj in gh_doc.Objects:
                if(obj.NickName == "MyToggle"):
                    obj.ExpireSolution(False)
                    
        # Define the WebView dialog class
        class WebviewSliderDialog(forms.Form):
            def __init__(self):
                super(WebviewSliderDialog, self).__init__()

                # Set up the WebView to load the HTML file
                self.Title = "Equivalent Frame Modelling"
                self.Padding = drawing.Padding(5)
                self.ClientSize = drawing.Size(600, 800)
                self.Resizable = True
                self.Topmost = True  # Keep the form on top

                # Create the WebView control and load the local HTML file
                self.m_webview = forms.WebView()
                self.m_webview.Size = drawing.Size(600, 800)
                self.m_webview.Url = System.Uri(path)  # Update with actual path
                

                # Subscribe to the DocumentLoading and DocumentLoaded events
                self.m_webview.DocumentLoaded += self.on_document_loaded
                self.m_webview.DocumentLoading += self.on_document_loading

                # Layout
                layout = forms.StackLayout()
                layout.Items.Add(forms.StackLayoutItem(self.m_webview, True))
                self.Content = layout
                self.set_center()

            # Get the Rhino main window handle and set the form to appear at the center of the screen
            def set_center(self):
                main_window = Rhino.UI.RhinoEtoApp.MainWindow
                screen_rect = main_window.Bounds
                self.Location = drawing.Point((screen_rect.Width - self.ClientSize.Width) // 2 + screen_rect.X,
                                              (screen_rect.Height - self.ClientSize.Height) // 2 + screen_rect.Y)
                
            def load_material_excel():
                try:
                    # Hard-coded path
                    excel_path = r"C:\Users\vrajasekar\Downloads\Materials.xlsx"

                    wb = openpyxl.load_workbook(excel_path, data_only=True)
                    sheet = wb[wb.sheetnames[0]]

                    field_names = []
                    rows_data = []
                    row_index = 0

                    for row in sheet.iter_rows(values_only=True):
                        if row_index == 0:
                            # First row = column headers
                            field_names = [str(x) for x in row if x is not None]
                        else:
                            row_dict = {}
                            # Populate row_dict by matching each cell to the correct field name
                            for i, col_name in enumerate(field_names):
                                row_dict[col_name] = row[i] if (i < len(row)) else None
                            rows_data.append(row_dict)
                        row_index += 1

                    # The shape your JS code expects
                    material_data = {
                        "fieldNames": field_names,
                        "rows": rows_data
                    }

                    # Store in sticky so we can retrieve it inside on_document_loaded
                    sc.sticky["my_material_data"] = material_data

                except Exception as ex:
                    print("Error reading Excel file:", ex)
                    sc.sticky["my_material_data"] = None
            
            load_material_excel() 
            
            # Inject JavaScript to handle initialization if needed
            def on_document_loaded(self, sender, e):
                try:
                    # Start building the js_script
                    js_script = "localStorage.clear();"

                    # 1) Elements injection
                    EFM = sc.sticky.get("EFM", {})
                    if "GeomDict" in EFM and "Elements" in EFM["GeomDict"]:
                        # Elements is now a list of dicts, e.g. [{ "name": ..., "thickness": ..., "geometries": ... }, ...]
                        elements_list = EFM["GeomDict"]["Elements"]
                    else:
                        elements_list = []

                    gh_doc.ExpireSolution()

                    # Convert the list of elements to JSON (but watch out if 'geometries' contains Rhino objects)
                    # For demonstration, let's ignore 'geometries' or store a placeholder
                    # A simple approach is to remove real geometry references before JSON:
                    cleaned_elements = []
                    for elem  in elements_list:
                        cleaned_elements.append({
                            "name": elem ["name"],
                            "thickness": elem ["thickness"],
                            # "geometries": ??? (can't directly JSON-serialize Rhino surfaces)
                            "geometries": []
                        })

                    elements_json = json.dumps(cleaned_elements, ensure_ascii=False)
                    js_script += f"""
                        var elementsData = {elements_json};
                        localStorage.setItem('elementsList', JSON.stringify(elementsData));
                        window.dispatchEvent(new Event('elementsLoaded'));
                    """
                    print("Executing js_script:", js_script)
                    sender.ExecuteScript(js_script)

                    # 2) Excel data injection
                    material_data = sc.sticky.get("my_material_data", None)
                    if material_data is not None:
                        # Convert to JSON
                        mat_json = json.dumps(material_data)
                        # Use the same pattern you do for geometry injection
                        js_script += f"""
                            window.materialData = {mat_json};
                            window.dispatchEvent(new Event('materialDataLoaded'));
                        """
                    else:
                        # If there's no data, we can at least set an empty structure
                        js_script += """
                            window.materialData = { fieldNames: [], rows: [] };
                            window.dispatchEvent(new Event('materialDataLoaded'));
                        """
                    
                    # 3) Stored materials from EFM["MatDict"]["Materials"]
                    mat_list = []
                    EFM = sc.sticky.get("EFM", {})
                    if "MatDict" in EFM and "Materials" in EFM["MatDict"]:
                        mat_list = EFM["MatDict"]["Materials"]
                    # Convert them to JSON. Might want to do a minimal "cleaning" if any invalid data
                    materials_json = json.dumps(mat_list, ensure_ascii=False)
                    js_script += f"""
                    // We'll store in localStorage or a global so the script can restore:
                    localStorage.setItem('materialsList', JSON.stringify({materials_json}));
                    window.dispatchEvent(new Event('materialsLoaded'));
                    """
                    print("Executing js_script:", js_script)
                    sender.ExecuteScript(js_script)

                except Exception as ex:
                    print("Error in on_document_loaded:", ex)    

            # Handle custom navigation events from JavaScript
            def on_document_loading(self, sender, e):

                if e.Uri.Scheme == "sliderupdate":
                    e.Cancel = True
                    value = e.Uri.Query.strip('?')
                    slider_id, slider_value = value.split('=')
                    sc.sticky[slider_id] = float(slider_value)
                    schedule_recompute()
                    refresh_sender()


                if e.Uri.Scheme == "geometryupdate":
                    e.Cancel = True  # Prevent actual navigation
                    value = e.Uri.Query.strip('?')
                    name_thickness = value.split(',')
                    if len(name_thickness) >= 2:
                        name = System.Uri.UnescapeDataString(name_thickness[0])
                        thickness = float(name_thickness[1])
                        surfaces = get_surfaces()

                        if surfaces:
                            # Access EFM from sticky
                            EFM = sc.sticky.get("EFM", {})
                            if "GeomDict" not in EFM:
                                EFM["GeomDict"] = {}
                            if "Elements" not in EFM["GeomDict"]:
                                EFM["GeomDict"]["Elements"] = []

                            # Try to find an existing element by name
                            found = False
                            for elem in EFM["GeomDict"]["Elements"]:
                                if elem["name"] == name:
                                    # Update thickness + geometry
                                    elem["thickness"] = thickness
                                    # Store actual Rhino surfaces in a python list
                                    elem["geometries"] = surfaces
                                    found = True
                                    break
                            if not found:
                                # Insert new element
                                new_elem = {
                                    "name": name,
                                    "thickness": thickness,
                                    "geometries": surfaces
                                }
                                EFM["GeomDict"]["Elements"].append(new_elem)

                            # Save back
                            sc.sticky["EFM"] = EFM

                            schedule_recompute()
                            refresh_sender()
                            self.BringToFront()  # bring form forward
                    else:
                        print("Invalid geometry update parameters.")

                if e.Uri.Scheme == "updateelements":
                    e.Cancel = True
                    try:
                        query = e.Uri.Query.strip('?')
                        params = dict(item.split('=') for item in query.split('&'))
                        encoded_data = params.get('data', '')

                        if encoded_data:
                            efm_json = System.Uri.UnescapeDataString(encoded_data)
                            new_efm_dict = json.loads(efm_json)  # This is what the browser sent
                            new_elements = new_efm_dict["GeomDict"]["Elements"]

                            # Get or create existing EFM in sticky
                            EFM = sc.sticky.get("EFM", {})

                            # If there's no GeomDict or Elements yet, initialize them
                            if "GeomDict" not in EFM:
                                EFM["GeomDict"] = {}
                            if "Elements" not in EFM["GeomDict"]:
                                EFM["GeomDict"]["Elements"] = []

                            old_elements = EFM["GeomDict"]["Elements"]

                            # Build a merged list
                            merged_list = []

                            for new_elem in new_elements:
                                # Find an old element that has the same name
                                matched_old = None
                                for old_el in old_elements:
                                    if old_el["name"] == new_elem["name"]:
                                        matched_old = old_el
                                        break

                                if matched_old:
                                    # Preserve old geometry, update thickness, etc.
                                    matched_old["thickness"] = new_elem["thickness"]
                                    # Optionally also update name if changed, or any other fields
                                    # BUT keep matched_old["geometries"] as is
                                    merged_list.append(matched_old)
                                else:
                                    # This is a brand-new element that didn't exist before.
                                    # Usually we won't have geometry yet; user must pick it
                                    # or you can store an empty list:
                                    new_elem["geometries"] = []
                                    merged_list.append(new_elem)

                            # Overwrite old list with the newly merged version
                            EFM["GeomDict"]["Elements"] = merged_list

                            # Store back
                            sc.sticky["EFM"] = EFM

                            # schedule a recompute so GH updates
                            schedule_recompute()
                            refresh_sender()

                    except Exception as ex:
                        print("Error updating elements (EFM):", ex)

                if e.Uri.Scheme == "deleteelement":
                    e.Cancel = True  # Prevent actual navigation
                    value = e.Uri.Query.strip('?')
                    params = dict(item.split('=') for item in value.split('&'))
                    name = System.Uri.UnescapeDataString(params.get('name', ''))
                    # if name:
                    #     remove_element_by_name(name)
                    #     schedule_recompute()
                    # else:
                    #     print("Invalid delete element parameters.")
                    if name:
                        # For demonstration, let's remove it from EFM if it exists
                        EFM = sc.sticky.get("EFM", {})
                        if "GeomDict" in EFM and "Elements" in EFM["GeomDict"]:
                            EFM["GeomDict"]["Elements"] = [
                                el for el in EFM["GeomDict"]["Elements"] if el["name"] != name
                            ]
                            sc.sticky["EFM"] = EFM
                        schedule_recompute()
                        refresh_sender()
                    else:
                        print("Invalid delete element parameters.")

                if e.Uri.Scheme == "updatematerials":
                    e.Cancel = True  # block actual navigation
                    try:
                        query = e.Uri.Query.strip('?')
                        params = dict(item.split('=') for item in query.split('&'))
                        encoded_data = params.get('data', '')

                        if encoded_data:
                            # 1) Parse JSON from the browser
                            mat_json = System.Uri.UnescapeDataString(encoded_data)
                            new_mat_dict = json.loads(mat_json)  

                            # 2) Get or create EFM["MatDict"]
                            EFM = sc.sticky.get("EFM", {})
                            if "MatDict" not in EFM:
                                EFM["MatDict"] = {}
                            if "Materials" not in EFM["MatDict"]:
                                EFM["MatDict"]["Materials"] = []

                            old_materials = EFM["MatDict"]["Materials"]
                            new_materials = new_mat_dict["MatDict"]["Materials"]

                            # 3) Merge them. For each new material, see if old one exists and update it
                            merged_list = []
                            for new_mat in new_materials:
                                matched_old = None
                                for old_mat in old_materials:
                                    if old_mat["name"] == new_mat["name"]:
                                        matched_old = old_mat
                                        break
                                if matched_old:
                                    # update existing
                                    for k, v in new_mat.items():
                                        matched_old[k] = v
                                    merged_list.append(matched_old)
                                else:
                                    # brand new
                                    merged_list.append(new_mat)

                            EFM["MatDict"]["Materials"] = merged_list
                            sc.sticky["EFM"] = EFM

                            # optionally:
                            schedule_recompute()
                            refresh_sender()

                    except Exception as ex:
                        print("Error updating materials (EFM):", ex)

                

        # Function to update elements list from table state
        def update_elements_from_table_state(table_state_json):
            table_state = json.loads(table_state_json)
            elements = sc.sticky.get("elements", [])

            for i, row_data in enumerate(table_state):
                name = row_data.get("name")
                thickness = float(row_data.get("thickness", 0))

                if len(elements) > i:
                    # Update existing element
                    elements[i].name = name
                    elements[i].thickness = thickness
                else:
                    # Add new element with empty geometries
                    new_element = Element(name, [], thickness)
                    elements.append(new_element)

            sc.sticky["elements"] = elements

        def remove_element_by_name(name):
            elements = sc.sticky.get('elements', [])
            # Remove the element with the matching name
            elements = [element for element in elements if element.name != name]
            sc.sticky['elements'] = elements

        # Function to add or update an element in the elements list
        def add_or_update_element(element):
            elements = sc.sticky.get("elements", [])

            # Check if an element with the same name already exists, and update it
            for i, existing_element in enumerate(elements):
                if existing_element.name == element.name:
                    elements[i] = element
                    sc.sticky["elements"] = elements
                    return

            # Otherwise, add the new element
            elements.append(element)
            sc.sticky["elements"] = elements

        # Run the UI within Rhino/Grasshopper based on a boolean variable
        if get_inputs:
            if 'form' not in sc.sticky or not isinstance(sc.sticky['form'], WebviewSliderDialog) or not sc.sticky['form'].Visible:
                form = WebviewSliderDialog()
                sc.sticky['form'] = form
                sc.sticky['form'].Show()
            else:
                # Bring the form to the front if it's already open
                sc.sticky['form'].BringToFront()
        
        # Return placeholders to GH
        return 
